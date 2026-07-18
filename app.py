"""
RongHui_SKU_HS_Check
独立于 EU Risk App / UK SKU Tracker 的系统，专门处理 RongHui（HONGKONG SICHANG
INTERNATIONAL GROUP）提供的 packing list，做 SKU 品名/HS 的一致性校验与主数据库维护。
"""
import io
import zipfile
from datetime import date

import pandas as pd
import streamlit as st

from pl_parser import parse_packing_list
from compare_engine import compare_batch, resolve_batch_conflicts, summarize
import sheets_db
from logo_const import VEVOR_LOGO_B64

st.set_page_config(page_title="RongHui_SKU_HS_Check", layout="wide")

APP_USERNAME = "Admin"
APP_PASSWORD = "Admin123"

CUSTOM_CSS = """
<style>
:root {
    --navy-deep: #0B1220;
    --navy-mid: #14233D;
    --brand-blue: #1E3A5F;
    --surface: #F7F8FA;
    --border-soft: #D8DEE4;
    --text-primary: #15202B;
    --text-muted: #5B6B7C;
}

/* 侧边栏：深墨蓝，企业系统感 */
section[data-testid="stSidebar"] {
    background-color: var(--navy-deep);
    border-right: 1px solid var(--navy-mid);
}
section[data-testid="stSidebar"] * {
    color: #E7ECF2 !important;
}
section[data-testid="stSidebar"] label, section[data-testid="stSidebar"] .stRadio label {
    color: #C7D2DD !important;
}

/* 侧边栏内按钮（如退出登录）背景浅色，文字需用深色才看得清 */
section[data-testid="stSidebar"] button {
    color: var(--text-primary) !important;
}
section[data-testid="stSidebar"] button p,
section[data-testid="stSidebar"] button span,
section[data-testid="stSidebar"] button div {
    color: var(--text-primary) !important;
}

/* 主内容区背景 */
.main .block-container {
    background-color: var(--surface);
    padding-top: 2rem;
}

/* 标题样式 */
h1, h2, h3 {
    color: var(--brand-blue);
    font-weight: 700;
    letter-spacing: -0.01em;
}

/* 指标卡片化 */
div[data-testid="stMetric"] {
    background-color: white;
    border: 1px solid var(--border-soft);
    border-radius: 6px;
    padding: 0.9rem 1rem;
}
div[data-testid="stMetric"] label {
    color: var(--text-muted) !important;
    font-weight: 600;
}

/* 信息/成功提示条，去掉默认圆润感，改为左侧色条风格 */
div[data-testid="stAlert"] {
    border-radius: 4px;
    border-left: 4px solid var(--brand-blue);
}

/* 表格容器边框 */
div[data-testid="stDataFrame"] {
    border: 1px solid var(--border-soft);
    border-radius: 6px;
}

/* 页脚签名 */
.app-footer {
    margin-top: 3rem;
    padding-top: 1rem;
    border-top: 1px solid var(--border-soft);
    text-align: center;
    color: var(--text-muted);
    font-size: 0.78rem;
    letter-spacing: 0.02em;
}
</style>
"""


def inject_global_style():
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


def render_sidebar_header():
    st.sidebar.markdown(
        f"""
        <div style="margin-bottom:1.2rem;">
            <img src="data:image/png;base64,{VEVOR_LOGO_B64}"
                 style="width:100%; border-radius:6px; display:block;" />
        </div>
        <hr style="border-color:#22324A; margin:0.8rem 0 1.2rem 0;">
        """,
        unsafe_allow_html=True,
    )


def render_sidebar_footer():
    st.sidebar.markdown(
        """
        <hr style="border-color:#22324A; margin:1.5rem 0 0.8rem 0;">
        <div style="color:#7C8CA0; font-size:0.78rem;">
            Designed and created by<br>
            <span style="color:#C7D2DD; font-weight:600;">Amélie — Vevor EU</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_footer():
    st.markdown(
        '<div class="app-footer">Designed and created by Amélie — Vevor EU</div>',
        unsafe_allow_html=True,
    )


def render_page_header(page_name: str):
    st.markdown(
        f"""
        <div style="margin-bottom:1.6rem;">
            <div style="font-size:2.1rem; font-weight:800; color:#1E3A5F; line-height:1.2;">
                RongHui_SKU_HS_Check
            </div>
            <div style="font-size:0.95rem; color:#5B6B7C; margin-top:0.3rem;">
                Vevor EU — Internal Use Only &nbsp;·&nbsp; {page_name}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def login_gate():
    inject_global_style()
    if st.session_state.get("authenticated"):
        return True
    st.title("RongHui_SKU_HS_Check")
    with st.form("login_form"):
        u = st.text_input("用户名")
        p = st.text_input("密码", type="password")
        submitted = st.form_submit_button("登录")
    if submitted:
        if u == APP_USERNAME and p == APP_PASSWORD:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("用户名或密码错误")
    render_footer()
    return False


def _fix_zip_filename_encoding(raw_name: str, flag_bits: int) -> str:
    """
    修复ZIP内中文文件名乱码：当ZIP条目未设置UTF-8标志位（flag_bits的0x800位）时，
    Python的zipfile库会按CP437解码文件名，但很多压缩工具（尤其Mac系统自带的"压缩"
    功能）实际写入的是UTF-8字节，导致中文文件名变成乱码。这里做一次"逆向"修复：
    把已经被错误解码成CP437的字符串重新编码回原始字节，再按UTF-8正确解码。
    """
    if flag_bits & 0x800:
        # 已经明确标记为UTF-8编码，说明zipfile已经解码正确，不需要处理
        return raw_name
    try:
        return raw_name.encode("cp437").decode("utf-8")
    except (UnicodeDecodeError, UnicodeEncodeError):
        # 修复失败（比如本来就是英文文件名，或者并非UTF-8字节），保留原始值
        return raw_name


def extract_excels_from_zip(zip_file) -> list[tuple[str, io.BytesIO]]:
    out = []
    with zipfile.ZipFile(zip_file) as zf:
        for info in zf.infolist():
            name = _fix_zip_filename_encoding(info.filename, info.flag_bits)
            if name.lower().endswith((".xlsx", ".xls")) and not name.startswith("__MACOSX"):
                data = zf.read(info.filename)
                out.append((name.split("/")[-1], io.BytesIO(data)))
    return out


def _reset_batch_state():
    for key in [
        "last_result_df", "last_file_count", "pending_conflicts", "pending_new",
        "new_skus_written", "today_new_skus", "batch_conflicts_resolved",
    ]:
        st.session_state.pop(key, None)


def _autosize_columns(ws, headers, export_df, min_width=12, max_width=80):
    """按每列实际内容最大长度自动设置列宽（留一点余量，并设置上限避免单列过宽）"""
    from openpyxl.utils import get_column_letter

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for value in export_df[header]:
            max_len = max(max_len, len(str(value)))
        width = min(max_len + 4, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, min_width)


def _build_simple_excel(export_df: pd.DataFrame, sheet_title: str) -> bytes:
    """通用Excel导出：表头加粗 + 按内容自动设置列宽，避免打开后内容被截断/看不全"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = list(export_df.columns)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(bold=True)

    for _, row in export_df.iterrows():
        ws.append([row[h] for h in headers])

    _autosize_columns(ws, headers, export_df)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_new_sku_summary_excel(export_df: pd.DataFrame) -> bytes:
    """生成当天新SKU总结表，自动设置足够宽的列宽，避免内容被截断"""
    return _build_simple_excel(export_df, "新SKU总结表")


def _build_batch_conflict_report(batch_conflicts: pd.DataFrame) -> bytes:
    """生成批次内部冲突Excel报告：同一SKU的所有变体行用浅蓝底色，不同SKU之间用粗黑分隔线"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "批次内部冲突"

    headers = ["序号", "冲突SKU", "版本", "品名", "HS编码", "税率", "出现次数(文件数)", "来源文件"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)

    light_blue = PatternFill("solid", start_color="DCEBFA", end_color="DCEBFA")
    thick_top = Border(top=Side(style="thick", color="333333"))

    row_idx = 2
    for n, (_, row) in enumerate(batch_conflicts.iterrows()):
        sku = row["sku"]
        variants = row["variants"]
        first_row_of_sku = row_idx
        for v_i, v in enumerate(variants):
            ws.append([
                n + 1, sku, v_i + 1, v["description"], v["hs_code"], v.get("tax_rate", ""),
                len(v["files"]), ", ".join(v["files"]),
            ])
            for c in range(1, len(headers) + 1):
                ws.cell(row_idx, c).fill = light_blue
            row_idx += 1
        # 在该SKU第一行的上边框加粗分隔线，区分不同SKU
        for c in range(1, len(headers) + 1):
            ws.cell(first_row_of_sku, c).border = thick_top

    widths = [6, 26, 6, 32, 14, 10, 16, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def page_upload_compare():
    render_page_header("上传与比对")
    upload_mode = st.radio("上传方式", ["ZIP批量上传", "单个/多个Excel文件"], horizontal=True)

    files_to_parse = []  # list of (filename, buffer)

    if upload_mode == "ZIP批量上传":
        zip_file = st.file_uploader("上传ZIP（包含多个packing list）", type=["zip"])
        if zip_file:
            files_to_parse = extract_excels_from_zip(zip_file)
            st.info(f"ZIP中识别到 {len(files_to_parse)} 个Excel文件")
    else:
        uploaded = st.file_uploader(
            "上传一个或多个packing list", type=["xlsx", "xls"], accept_multiple_files=True
        )
        if uploaded:
            files_to_parse = [(f.name, f) for f in uploaded]

    if not files_to_parse:
        return

    if st.button("开始解析与比对", type="primary"):
        _reset_batch_state()
        all_records = []
        parse_errors = []
        progress = st.progress(0.0)
        for i, (fname, buf) in enumerate(files_to_parse):
            try:
                df = parse_packing_list(buf, fname)
                all_records.append(df)
            except Exception as e:
                parse_errors.append((fname, str(e)))
            progress.progress((i + 1) / len(files_to_parse))

        if parse_errors:
            with st.expander(f"⚠️ {len(parse_errors)} 个文件解析失败", expanded=True):
                for fname, err in parse_errors:
                    st.error(f"{fname}: {err}")

        if not all_records:
            st.warning("没有成功解析的文件")
            return

        new_df = pd.concat(all_records, ignore_index=True)
        db_df = sheets_db.load_master_db()
        result_df = compare_batch(new_df, db_df)

        st.session_state["last_result_df"] = result_df
        st.session_state["last_file_count"] = len(files_to_parse) - len(parse_errors)
        st.session_state["last_db_df"] = db_df

    if "last_result_df" not in st.session_state:
        return

    result_df = st.session_state["last_result_df"]

    # ---------- 第一步：批次内部冲突必须先处理 ----------
    batch_conflicts = result_df[result_df["status"] == "BATCH_CONFLICT"]
    if len(batch_conflicts) > 0 and not st.session_state.get("batch_conflicts_resolved", False):
        st.subheader(f"⚠️ 批次内部冲突（{len(batch_conflicts)} 个SKU）")
        st.caption(
            "以下SKU在本次上传的不同文件里，品名或HS不一致。请为每个SKU选择本次最终采用哪个版本"
            "（默认已预选「出现文件数最多」的版本，可在下拉框里修改）。处理完才能继续后续比对。"
        )

        report_buf = _build_batch_conflict_report(batch_conflicts)
        st.download_button(
            "📥 一键下载批次内部冲突报告 (Excel)",
            data=report_buf,
            file_name=f"RongHui_BatchConflicts_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.caption("👉 每条记录请勾选左侧「已确认」复选框，表示你已经看过并确认这条的处理方式。")

        choice_rows = []
        for _, row in batch_conflicts.iterrows():
            variants = row["variants"]
            default_idx = max(range(len(variants)), key=lambda i: len(variants[i]["files"]))
            options = [
                f"[{i+1}] {v['description']} | HS={v['hs_code']} | 税率={v.get('tax_rate', '') or '（空）'} "
                f"| 来自{len(v['files'])}个文件: {', '.join(v['files'])}"
                for i, v in enumerate(variants)
            ]
            choice_rows.append({
                "SKU": row["sku"],
                "选择采用哪个版本": options[default_idx],
                "_options": options,
                "_sku": row["sku"],
                "_default_label": options[default_idx],
            })

        scroll_box = st.container(height=520, border=True)
        with scroll_box:
            for n, item in enumerate(choice_rows):
                col_check, col_select = st.columns([1, 9])
                with col_check:
                    st.markdown("<div style='margin-top:1.8rem;'></div>", unsafe_allow_html=True)
                    checked = st.checkbox(
                        "已确认", key=f"batch_conflict_checked_{item['_sku']}", label_visibility="collapsed"
                    )
                with col_select:
                    choice_rows[n]["选择采用哪个版本"] = st.selectbox(
                        f"{n + 1}. SKU: {item['SKU']}",
                        options=item["_options"],
                        index=item["_options"].index(item["选择采用哪个版本"]),
                        key=f"batch_conflict_{item['_sku']}",
                    )
                choice_rows[n]["_checked"] = checked

        # 统计有多少SKU还没被人工勾选「已确认」
        unchecked_count = sum(1 for item in choice_rows if not item["_checked"])

        def _do_confirm():
            resolutions = {}
            for item in choice_rows:
                chosen_label = item["选择采用哪个版本"]
                chosen_idx = item["_options"].index(chosen_label)
                resolutions[item["_sku"]] = chosen_idx

            resolved_df = resolve_batch_conflicts(result_df, resolutions, st.session_state["last_db_df"])
            st.session_state["last_result_df"] = resolved_df
            st.session_state["batch_conflicts_resolved"] = True
            st.session_state["pending_conflicts"] = resolved_df[resolved_df["status"] == "CONFLICT"].copy()
            st.session_state["pending_new"] = resolved_df[resolved_df["status"] == "NEW"].copy()
            st.session_state.pop("show_batch_confirm_warning", None)
            st.rerun()

        if st.session_state.get("show_batch_confirm_warning", False):
            st.warning(
                f"还有 {unchecked_count} 个SKU你尚未勾选「已确认」，尚未逐条检查确认。"
                "是否仍要继续，直接采用当前下拉框里的值（含未检查项的默认推荐值）？"
            )
            c1, c2 = st.columns(2)
            with c1:
                if st.button("仍然继续（采用当前值）", type="primary"):
                    _do_confirm()
            with c2:
                if st.button("我再检查一下"):
                    st.session_state.pop("show_batch_confirm_warning", None)
                    st.rerun()
        else:
            if st.button("✅ 确认批次内部冲突的选择，继续比对", type="primary"):
                if unchecked_count > 0:
                    st.session_state["show_batch_confirm_warning"] = True
                    st.rerun()
                else:
                    _do_confirm()
        return  # 批次内部冲突没处理完，不展示后续报告

    if len(batch_conflicts) == 0:
        st.session_state.setdefault("pending_conflicts", result_df[result_df["status"] == "CONFLICT"].copy())
        st.session_state.setdefault("pending_new", result_df[result_df["status"] == "NEW"].copy())
        st.session_state.setdefault("batch_conflicts_resolved", True)

    result_df = st.session_state["last_result_df"]
    stats = summarize(result_df, st.session_state["last_file_count"])

    st.subheader("本次录入报告")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("录入文件数", stats["file_count"])
    c2.metric("合计SKU数", stats["total_skus"])
    c3.metric("已存在且一致", stats["match_count"])
    c4.metric("新SKU", stats["new_count"])
    c5.metric("需确认冲突", stats["conflict_count"])

    summary_export_df = result_df[["sku", "new_description", "new_hs_code", "new_tax_rate", "status", "sources"]].rename(
        columns={"new_description": "description", "new_hs_code": "hs_code", "new_tax_rate": "tax_rate"}
    )
    summary_buf = _build_simple_excel(summary_export_df, "本次核查总结")
    st.download_button(
        "下载本次核查总结单 (Excel)",
        data=summary_buf,
        file_name=f"RongHui_Check_Summary_{date.today().isoformat()}_{stats['file_count']}containers.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_check_summary",
    )

    pending_new = st.session_state.get("pending_new", pd.DataFrame())
    if len(pending_new) > 0 and not st.session_state.get("new_skus_written", False):
        st.info(f"本次发现 {len(pending_new)} 个新SKU，预览如下，确认无误后点击下方按钮写入数据库")
        preview_search = st.text_input("搜索新SKU预览（按SKU或品名）", key="preview_search")
        preview_df = pending_new[["sku", "new_description", "new_hs_code", "new_tax_rate", "sources"]].rename(
            columns={"new_description": "品名", "new_hs_code": "HS编码", "new_tax_rate": "税率", "sku": "SKU", "sources": "来源文件"}
        )
        if preview_search:
            mask = (
                preview_df["SKU"].str.contains(preview_search, case=False, na=False)
                | preview_df["品名"].str.contains(preview_search, case=False, na=False)
            )
            preview_df = preview_df[mask]
        st.dataframe(preview_df, use_container_width=True, hide_index=True, height=350)

        if st.button("✅ 确认写入新SKU到数据库", type="primary"):
            new_rows = [
                {
                    "sku": r["sku"], "description": r["new_description"],
                    "hs_code": r["new_hs_code"], "tax_rate": r["new_tax_rate"],
                    "sources": r["sources"],
                }
                for _, r in pending_new.iterrows()
            ]
            sheets_db.append_new_skus(new_rows)
            st.session_state["new_skus_written"] = True
            st.session_state["today_new_skus"] = pending_new
            st.success(f"已写入 {len(new_rows)} 个新SKU到数据库")
            st.rerun()

    pending_conflicts = st.session_state.get("pending_conflicts", pd.DataFrame())
    if len(pending_conflicts) > 0:
        st.subheader(f"需人工确认的数据库冲突（{len(pending_conflicts)} 个）")
        st.caption("勾选「更新为新记录」的SKU将用本次packing list的数据覆盖数据库；未勾选的默认保留数据库原记录")

        display_df = pending_conflicts.copy()
        display_df.insert(0, "更新为新记录", False)
        edited = st.data_editor(
            display_df[[
                "更新为新记录", "sku", "db_description", "db_hs_code", "db_tax_rate",
                "new_description", "new_hs_code", "new_tax_rate", "sources",
            ]],
            column_config={
                "更新为新记录": st.column_config.CheckboxColumn(),
                "sku": "SKU",
                "db_description": "数据库原品名",
                "db_hs_code": "数据库原HS",
                "db_tax_rate": "数据库原税率",
                "new_description": "本次品名",
                "new_hs_code": "本次HS",
                "new_tax_rate": "本次税率",
                "sources": "来源文件",
            },
            disabled=[
                "sku", "db_description", "db_hs_code", "db_tax_rate",
                "new_description", "new_hs_code", "new_tax_rate", "sources",
            ],
            hide_index=True,
            use_container_width=True,
            key="conflict_editor",
        )
        st.caption("说明：是否需要更新只看品名/HS是否一致，税率仅作展示；勾选「更新为新记录」时税率也会一并覆盖为本次的值。")

        if st.button("提交数据库冲突处理结果", type="primary"):
            decisions = []
            for _, row in edited.iterrows():
                orig = pending_conflicts[pending_conflicts["sku"] == row["sku"]].iloc[0]
                decisions.append({
                    "sku": row["sku"],
                    "action": "UPDATE" if row["更新为新记录"] else "KEEP",
                    "old_description": orig["db_description"],
                    "old_hs_code": orig["db_hs_code"],
                    "old_tax_rate": orig["db_tax_rate"],
                    "new_description": orig["new_description"],
                    "new_hs_code": orig["new_hs_code"],
                    "new_tax_rate": orig["new_tax_rate"],
                })
            sheets_db.apply_conflict_decisions(decisions)
            update_count = sum(1 for d in decisions if d["action"] == "UPDATE")
            keep_count = len(decisions) - update_count
            st.success(f"已处理 {len(decisions)} 个冲突：{update_count} 个更新，{keep_count} 个保留原记录")
            st.session_state.pop("pending_conflicts", None)
    elif st.session_state.get("batch_conflicts_resolved", False):
        st.info("本次没有需要人工确认的数据库冲突")

    today_new = st.session_state.get("today_new_skus")
    if today_new is not None and len(today_new) > 0:
        st.subheader("当天新SKU总结表")
        export_df = today_new[["sku", "new_description", "new_hs_code", "new_tax_rate", "sources"]].rename(
            columns={"new_description": "description", "new_hs_code": "hs_code", "new_tax_rate": "tax_rate"}
        )
        st.dataframe(export_df, use_container_width=True, hide_index=True)

        container_count = st.session_state.get("last_file_count", len(today_new))
        buf = _build_new_sku_summary_excel(export_df)
        st.download_button(
            "下载当天新SKU总结表 (Excel)",
            data=buf,
            file_name=f"RongHui_New_SKUs_{date.today().isoformat()}_{container_count}containers.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def page_database():
    render_page_header("主数据库")
    db_df = sheets_db.load_master_db()
    st.caption(f"共 {len(db_df)} 条SKU记录")

    search = st.text_input("搜索SKU或品名")
    display_df = db_df
    if search:
        mask = db_df["sku"].str.contains(search, case=False, na=False) | db_df["description"].str.contains(search, case=False, na=False)
        display_df = db_df[mask]

    st.dataframe(display_df, use_container_width=True, hide_index=True)

    buf = _build_simple_excel(db_df, "主数据库")
    st.download_button(
        "一键下载主数据库 (Excel)",
        data=buf,
        file_name=f"RongHui_SKU_Master_{date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def page_change_log():
    render_page_header("变更日志")
    log_df = sheets_db.load_change_log()
    st.caption(f"共 {len(log_df)} 条记录")
    action_filter = st.multiselect("筛选操作类型", options=sorted(log_df["action"].unique()) if len(log_df) else [])
    display_df = log_df
    if action_filter:
        display_df = log_df[log_df["action"].isin(action_filter)]
    st.dataframe(display_df.sort_values("timestamp", ascending=False) if len(display_df) else display_df,
                 use_container_width=True, hide_index=True)


def main():
    if not login_gate():
        return

    render_sidebar_header()
    page = st.sidebar.radio("导航", ["上传与比对", "主数据库", "变更日志"])
    if st.sidebar.button("退出登录"):
        st.session_state["authenticated"] = False
        st.rerun()
    render_sidebar_footer()

    if page == "上传与比对":
        page_upload_compare()
    elif page == "主数据库":
        page_database()
    elif page == "变更日志":
        page_change_log()


if __name__ == "__main__":
    main()
