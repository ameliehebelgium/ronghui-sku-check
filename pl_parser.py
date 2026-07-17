"""
RongHui Packing List 解析模块

解析逻辑：
0. 文件可能包含多个分单（发票/箱单/合同等），固定读取名称中包含"箱单"的那个
   sheet，不再默认取第一个 sheet。
1. 先定位两个一定带文字的"锚点列"：
   - "Item Description" 列（描述兜底列）
   - "HTS" 列（HS兜底列）
2. 按位置关系推断"优先列"（不依赖表头文字是否存在，因为部分批次这两列没有标题）：
   - "建议品名" = Item Description 右边紧邻一列（前提：该列不是 HTS 本身，且不是 SKU 列）
   - "建议HS"   = HTS 右边紧邻一列（前提：该列不是下一个已知字段列，比如 Qty/税率，且不是 SKU 列）
   如果该位置紧邻的就是另一个已知字段（比如直接是 Qty / 税率），说明这次文件没有"建议"列，
   优先列按 None 处理，自动回退用兜底列。
   - "税率"（本次新增） = 建议HS 右边紧邻一列。只有在"建议HS"这一列存在时才去抓取，
     因为这一列本身就是跟着"建议HS"出现的，不单独设兜底列。
3. 描述列/HS列/税率列都做 forward-fill（合并格/续行导致的空值，向上找最近非空值），
   三者共用同一套"是否进入新分组"的判断（以描述兜底列出现新值为准）。
4. SKU 列没有固定位置/表头，用内容特征自动识别（每行几乎都非空、且为字母数字混合长字符串）。
5. 排除 TOTAL / Signature 等汇总行（SKU列为空或不符合SKU模式的行自动跳过）。
6. 税率原始写法不统一（"0.027" / "0,027" / "2,7%" 等），统一格式化成"2,7%"这种
   "逗号小数 + 百分号"的形式后再输出。
"""
import re
import pandas as pd
import openpyxl


DESC_FALLBACK_KEYWORDS = ["item description", "description"]
HS_FALLBACK_KEYWORDS = ["hts", "hs code", "h.s.code", "h.s. code"]
NO_KEYWORDS = ["no.", "no", "序号"]
# 已知的其它字段表头关键字，用来判断"右边紧邻列"是否其实是别的已知字段
# （也就是说这次文件没有"建议品名/建议HS"这一列）
KNOWN_OTHER_FIELD_KEYWORDS = [
    "qty", "ctns", "g.w.", "n.w.", "税率", "hts", "hs code",
]

SKU_PATTERN = re.compile(r"^[A-Z0-9][A-Z0-9\-]{6,}$")


def _select_sheet(wb, file_name=None):
    """文件可能含多个分单（发票/箱单/合同...），固定选择名称里带"箱单"的那个"""
    matches = [name for name in wb.sheetnames if "箱单" in name]
    if not matches:
        raise ValueError(
            f"找不到'箱单'分单（{file_name}）：工作簿里的分单是 {wb.sheetnames}"
        )
    return wb[matches[0]]


def _find_header_row(ws, max_scan_rows=40):
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower().rstrip(".") in [
                k.rstrip(".") for k in NO_KEYWORDS
            ]:
                return r
    return None


def _match_col(ws, header_row, keywords):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str):
            vv = v.strip().lower()
            for kw in keywords:
                if kw in vv:
                    return c
    return None


def _header_text(ws, header_row, col):
    v = ws.cell(header_row, col).value
    return v.strip().lower() if isinstance(v, str) else None


def _is_known_other_field(ws, header_row, col):
    """判断某列的表头是否是一个已知的、与“建议XX”无关的字段名"""
    txt = _header_text(ws, header_row, col)
    if txt is None:
        return False
    return any(kw in txt for kw in KNOWN_OTHER_FIELD_KEYWORDS)


def _looks_like_sku(value):
    if value is None:
        return False
    s = str(value).strip()
    if not s or s.replace(".", "", 1).isdigit():
        return False
    return bool(SKU_PATTERN.match(s.upper()))


def _find_sku_col(ws, header_row, exclude_cols):
    data_start = header_row + 1
    data_end = ws.max_row
    best_col, best_score = None, 0
    for c in range(1, ws.max_column + 1):
        if c in exclude_cols:
            continue
        total, hits = 0, 0
        for r in range(data_start, data_end + 1):
            v = ws.cell(r, c).value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            total += 1
            if _looks_like_sku(v):
                hits += 1
        if total >= 3 and hits / total > 0.8 and total > best_score:
            best_score = total
            best_col = c
    return best_col


def _find_preferred_col(ws, header_row, anchor_col, sku_col_guess_exclude):
    """
    给定一个锚点列（Item Description 或 HTS），返回其右边紧邻一列的列号，
    前提：右边紧邻列的表头不是另一个已知字段（说明这次文件确实没有"建议"列）。
    如果锚点列本身不存在，返回 None。
    """
    if anchor_col is None:
        return None
    candidate = anchor_col + 1
    if candidate in sku_col_guess_exclude:
        return None
    if _is_known_other_field(ws, header_row, candidate):
        return None
    return candidate


def parse_packing_list(file_path_or_buffer, file_name=None):
    """
    解析单个 packing list 文件，返回 DataFrame:
    columns = [sku, description, hs_code, tax_rate, po_number, source_file]
    """
    wb = openpyxl.load_workbook(file_path_or_buffer, data_only=True)
    ws = _select_sheet(wb, file_name)

    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError(f"找不到表头行（{file_name}）：未发现 'No.' 类关键字")

    desc_fallback_col = _match_col(ws, header_row, DESC_FALLBACK_KEYWORDS)
    hs_fallback_col = _match_col(ws, header_row, HS_FALLBACK_KEYWORDS)

    if desc_fallback_col is None:
        raise ValueError(f"找不到描述列（{file_name}）：未发现 'Item Description'")
    if hs_fallback_col is None:
        raise ValueError(f"找不到HS列（{file_name}）：未发现 'HTS'")

    # 建议品名/建议HS：按位置关系推断（右边紧邻列），不依赖该列是否有表头文字
    desc_primary_col = _find_preferred_col(ws, header_row, desc_fallback_col, set())
    hs_primary_col = _find_preferred_col(ws, header_row, hs_fallback_col, set())

    # 税率列：紧跟在"建议HS"右边一列，只有"建议HS"这一列存在时才有意义
    tax_col = (hs_primary_col + 1) if hs_primary_col else None

    exclude = {c for c in [desc_primary_col, desc_fallback_col, hs_primary_col, hs_fallback_col, tax_col] if c}
    sku_col = _find_sku_col(ws, header_row, exclude)
    if sku_col is None:
        raise ValueError(f"找不到SKU列（{file_name}）：内容特征识别失败")

    po_number = None
    for r in range(1, header_row):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "po no" in v.strip().lower():
                po_number = ws.cell(r, c + 1).value
                break
        if po_number:
            break

    records = []
    last_desc_primary, last_desc_fallback = None, None
    last_hs_primary, last_hs_fallback = None, None
    last_tax = None

    for r in range(header_row + 1, ws.max_row + 1):
        df_raw = ws.cell(r, desc_fallback_col).value
        is_new_group = df_raw is not None and str(df_raw).strip() != ""

        sku_val = ws.cell(r, sku_col).value
        dp = ws.cell(r, desc_primary_col).value if desc_primary_col else None
        hp = ws.cell(r, hs_primary_col).value if hs_primary_col else None
        hf = ws.cell(r, hs_fallback_col).value if hs_fallback_col else None
        tx = ws.cell(r, tax_col).value if tax_col else None

        dp_str = str(dp).strip() if dp is not None else ""
        df_str = str(df_raw).strip() if df_raw is not None else ""
        hp_str = str(hp).strip() if hp is not None else ""
        hf_str = str(hf).strip() if hf is not None else ""
        tx_str = str(tx).strip() if tx is not None else ""

        if is_new_group:
            # fallback 列出现新分组值：更新 fallback 缓存，并重置 primary 缓存。
            # primary 缓存只在「与 fallback 同行出现」时才作为组级继承值；
            # 这样子行里单独出现的 primary 值（如 Exhaust system）只服务当行，
            # 后续空行会正确回退到 fallback（如 DIESEL HEATER PARTS），而不是
            # 继续沿用子行的 primary。
            last_desc_fallback = df_str
            last_desc_primary = dp_str  # 同行 primary 有值就记录，没有就清空
            last_hs_fallback = hf_str if hf_str else last_hs_fallback
            last_hs_primary = hp_str  # 同上
            last_tax = tx_str  # 税率跟 hs_primary 同一套规则：同行有值就记录，没有就清空
        else:
            # 非新分组（子行）：desc_primary / hs_primary / 税率都做 forward-fill。
            # Excel 合并格在 openpyxl 里表现为"只有第一格有值，后续格为 None"，
            # None 即"继承上一格"，forward-fill 正好还原这个语义。
            # 唯一重置时机：is_new_group（fallback 列出现新值，进入新产品组）。
            if df_str:
                last_desc_fallback = df_str
            if dp_str:
                last_desc_primary = dp_str
            if hf_str:
                last_hs_fallback = hf_str
            if hp_str:
                last_hs_primary = hp_str
            if tx_str:
                last_tax = tx_str

        if not _looks_like_sku(sku_val):
            continue

        # primary 优先，都没有才回退到 fallback
        description = last_desc_primary if last_desc_primary else last_desc_fallback
        hs_code = last_hs_primary if last_hs_primary else last_hs_fallback

        records.append({
            "sku": str(sku_val).strip(),
            "description": str(description).strip() if description is not None else "",
            "hs_code": _normalize_hs(hs_code),
            "tax_rate": _normalize_tax_rate(last_tax),
            "po_number": str(po_number).strip() if po_number else "",
            "source_file": file_name or "",
        })

    return pd.DataFrame(records)


def _normalize_tax_rate(value):
    """
    统一税率写法为"2,7%"（逗号小数 + 百分号）：
    - 带 % 号的写法（如 "2,7%" / "2.7%"）：% 前面的数字本身就是百分比数值
    - 不带 % 号的写法（如 "0,027" / "0.027" / 数字 0.027）：视为小数比例，乘以100换算成百分比数值
    换算后统一用逗号做小数分隔符，去掉多余的小数位（整数时不带小数点）。
    """
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""

    is_percent = "%" in s
    s_clean = s.replace("%", "").replace(",", ".").strip()
    if not s_clean:
        return ""

    try:
        num = float(s_clean)
    except ValueError:
        # 无法识别的写法，原样返回，避免静默丢数据
        return s

    if not is_percent:
        num *= 100
    num = round(num, 4)

    if num == int(num):
        formatted = str(int(num))
    else:
        formatted = f"{num:.4f}".rstrip("0").rstrip(".").replace(".", ",")

    return f"{formatted}%"


def _normalize_hs(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s
